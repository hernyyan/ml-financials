"""
Temporary script — builds the two-tab reconciliation workbook comparing the
confirmed current monitoring template (47 items) against every other Fabric
data_item filed under Income Statement / Balance Sheet / Cash Flow, with a
proposed MATCH or ABANDON call per row for the user to review/override.
"""
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "exploration")
OUT_PATH = os.path.join(os.path.expanduser("~"), "Downloads", "fabric_line_item_reconciliation.xlsx")

# (data_item_id, statement, section) for the 47 confirmed current template items.
# Order matches the confirmed monitoring-workbook template.
CONFIRMED = [
    (10717, "Income Statement", ""), (10001, "Income Statement", ""),
    (11679, "Income Statement", ""), (11680, "Income Statement", ""),
    (11682, "Income Statement", ""), (10059, "Income Statement", ""),
    (11705, "Income Statement", ""), (10004, "Income Statement", ""),
    (10012, "Income Statement", ""), (12190, "Income Statement", ""),
    (10014, "Income Statement", ""), (11683, "Income Statement", ""),
    (12212, "Income Statement", "LTM - Adj EBITDA Items"),
    (12211, "Income Statement", "LTM - Adj EBITDA Items"),
    (10722, "Income Statement", "LTM - Adj EBITDA Items"),
    (10016, "Balance Sheet", "Assets"), (10017, "Balance Sheet", "Assets"),
    (10269, "Balance Sheet", "Assets"), (10018, "Balance Sheet", "Assets"),
    (10020, "Balance Sheet", "Assets"), (11684, "Balance Sheet", "Assets"),
    (10270, "Balance Sheet", "Assets"), (10271, "Balance Sheet", "Assets"),
    (10023, "Balance Sheet", "Assets"), (10025, "Balance Sheet", "Assets"),
    (11686, "Balance Sheet", "Assets"), (11687, "Balance Sheet", "Assets"),
    (10028, "Balance Sheet", "Liabilities"), (10030, "Balance Sheet", "Liabilities"),
    (10031, "Balance Sheet", "Liabilities"), (10730, "Balance Sheet", "Liabilities"),
    (12179, "Balance Sheet", "Liabilities"), (10032, "Balance Sheet", "Liabilities"),
    (12192, "Balance Sheet", "Liabilities"), (10728, "Balance Sheet", "Liabilities"),
    (10727, "Balance Sheet", "Liabilities"), (12191, "Balance Sheet", "Liabilities"),
    (12194, "Balance Sheet", "Liabilities"), (11689, "Balance Sheet", "Liabilities"),
    (10040, "Balance Sheet", "Equity"), (10041, "Balance Sheet", "Equity"),
    (12193, "Balance Sheet", "Equity"), (11690, "Balance Sheet", "Equity"),
    (10047, "Cash Flow", ""), (10048, "Cash Flow", ""),
    (10049, "Cash Flow", ""), (10292, "Cash Flow", ""),
]
CONFIRMED_IDS = {row[0] for row in CONFIRMED}
CONFIRMED_META = {row[0]: row for row in CONFIRMED}

# Confirmed retired -> current renames (name-suffix "- OLD DO NOT USE" match).
RENAME_MAP = {
    10033: 12192,  # Total Current Liabilities - OLD DO NOT USE -> Total Current Liabilities
    10063: 12194,  # Total Non-Current Liabilities - OLD DO NOT USE -> Total Non-Current Liabilities
    10036: 12191,  # Other Non-Current Liabilities - OLD DO NOT USE -> Other Non-Current Liabilities
}

# Retired items confirmed to have no live replacement.
RETIRED_NO_REPLACEMENT = {10724, 10725, 10729}

TEST_JUNK_IDS = {11554, 11555, 11556}
MISCATEGORIZED_IDS = {11288, 11289, 11287}  # End Market % / Explanation / Name


def load_all_items():
    with open(os.path.join(DATA_DIR, "financial_statement_items_all.csv"), encoding="utf-8") as f:
        return list(csv.DictReader(f))


def classify(item):
    """Return (proposed_action, notes) for a non-confirmed item."""
    item_id = int(item["data_item_id"])
    name = item["data_item_name"]

    if item_id in RENAME_MAP:
        target_id = RENAME_MAP[item_id]
        target_name = next(i["data_item_name"] for i in ALL_ITEMS if int(i["data_item_id"]) == target_id) \
            if any(int(i["data_item_id"]) == target_id for i in ALL_ITEMS) else "?"
        return f"MATCH -> {target_name} ({target_id})", "Retired rename, name-suffix confirmed match"

    if item_id in RETIRED_NO_REPLACEMENT:
        return "ABANDON", "Retired (OLD DO NOT USE), no live replacement found"

    if item_id in TEST_JUNK_IDS:
        return "ABANDON", "Test/junk data item, not real"

    if item_id in MISCATEGORIZED_IDS:
        return "ABANDON", "Miscategorized under this statement; not a real financial line"

    if "OLD DO NOT USE" in name:
        return "ABANDON", "Retired, not otherwise resolved -- confirm no replacement needed"

    if name.endswith("_calc"):
        return "ABANDON", "System-calculated shadow/check value, not source data"

    if "Margin %" in name or "/EBITDA" in name:
        return "ABANDON", "Ratio/margin metric -- bronze stores raw dollars only"

    if item_id in (10038, 10273):
        return f"MATCH -> Paid In Capital (10040)?", "Judgment call: fold Common/Preferred Stock into Paid in Capital, or keep separate?"

    if item_id == 10034:
        return "MATCH -> Long Term Loans (10728)?", "Judgment call: is Long Term Debt the same as Long Term Loans, or a broader bucket incl. leases?"

    if item_id in (10275, 10276):
        return "ABANDON?", "Was in prior (superseded) schema but not on confirmed template -- keep as extra bronze column anyway?"

    if item_id in (10715, 10716):
        return "ABANDON", "Template only tracks Total Revenue; this is a distinct/unused revenue variant"

    if item_id == 12188:
        return "ABANDON?", "Live (not OLD-tagged) item very similar to Total Operating Expenses (11680) -- possibly a genuine duplicate/rename not caught by the OLD-tag convention"

    if item_id == 10721:
        return "ABANDON?", "Live (not OLD-tagged) item similar to Other Expense/(Income) (12190) -- possibly predecessor not caught by OLD-tag convention"

    if item_id in (11681, 10535):
        return "ABANDON", "Total Expense/(Income) not on confirmed template (component-level Other Expense/(Income) is used instead)"

    if item_id == 12214 or item_id == 10044:
        return "ABANDON", "Total Liabilities and Equity is a balance-check row, derived not stored"

    if item_id == 10723:
        return "ABANDON", "Net Operating Income was only in prior superseded schema, not on confirmed template"

    return "ABANDON", "Granular sub-line item not on the confirmed rollup template (scope decision)"


def build_workbook():
    global ALL_ITEMS
    ALL_ITEMS = load_all_items()

    wb = Workbook()

    # --- Tab 1: confirmed current template ---
    ws1 = wb.active
    ws1.title = "Confirmed Template Items"
    header = ["data_item_id", "data_item_name", "statement", "section"]
    ws1.append(header)
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    by_id = {int(i["data_item_id"]): i for i in ALL_ITEMS}
    for item_id, statement, section in CONFIRMED:
        name = by_id[item_id]["data_item_name"] if item_id in by_id else "(not found)"
        ws1.append([item_id, name, statement, section])

    # --- Tab 2: everything else, with proposed action ---
    ws2 = wb.create_sheet("Other Fabric FS Items")
    header2 = ["data_item_id", "data_item_name", "category", "proposed_action", "notes", "your_decision"]
    ws2.append(header2)
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    review_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for item in ALL_ITEMS:
        item_id = int(item["data_item_id"])
        if item_id in CONFIRMED_IDS:
            continue
        action, notes = classify(item)
        row = [item_id, item["data_item_name"], item["data_item_category"], action, notes, ""]
        ws2.append(row)
        if "?" in action:
            for col in range(1, 7):
                ws2.cell(row=ws2.max_row, column=col).fill = review_fill

    for ws in (ws1, ws2):
        for col_idx in range(1, ws.max_column + 1):
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"Tab 1 (confirmed): {len(CONFIRMED)} rows")
    print(f"Tab 2 (other): {ws2.max_row - 1} rows")


if __name__ == "__main__":
    build_workbook()
