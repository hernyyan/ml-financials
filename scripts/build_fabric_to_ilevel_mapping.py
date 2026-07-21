"""
Temporary script -- builds an organized, visual workbook documenting the
Fabric -> iLEVEL API migration: connection/auth differences, resource-level
mapping (which iLEVEL endpoint replaces which Fabric table), and field-level
mapping for the two tables that matter most (data_items, periodic_data).
Grounded in a live pull against the iLEVEL API on 2026-07-20, not the
Postman docs alone.
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_PATH = os.path.join(os.path.expanduser("~"), "Downloads", "fabric_to_ilevel_mapping.xlsx")

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FLAG_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
NEW_FIELD_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
PII_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def autosize(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_rows(ws, rows, fills=None):
    for i, row in enumerate(rows):
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(row) + 1):
            ws.cell(row=r, column=c).alignment = WRAP
        if fills and fills[i]:
            for c in range(1, len(row) + 1):
                ws.cell(row=r, column=c).fill = fills[i]


def build():
    wb = Workbook()

    # --- Tab 1: Overview ---
    ws = wb.active
    ws.title = "Overview"
    ws.append(["Topic", "Detail"])
    style_header(ws, 2)
    rows = [
        ["Migration",
         "Replace the Fabric SQL Warehouse (smc_dwh, pulled via pyodbc + Azure AD service principal) "
         "with direct calls to the iLEVEL REST API as the source for portfolio company financial data."],
        ["Key finding (confirmed 2026-07-20)",
         "iLEVEL is the literal system of record that Fabric's production.base_ilevel__* tables are "
         "warehoused from. A live pull of GET /dataItems returned all 1,683 data items, and the ids/names "
         "matched Fabric's base_ilevel__data_items exactly -- including the same 6 retired "
         "\"- OLD DO NOT USE\" items (10729, 10036, 10724, 10725, 10033, 10063)."],
        ["Correction to prior assumption",
         "iLEVEL does NOT expose only current/clean line items. /dataItems returns the full lifetime "
         "catalog, same as Fabric. This means the already-completed 47-item confirmed template "
         "reconciliation and the OLD-items volume investigation (fabric_line_item_reconciliation.xlsx) "
         "remain fully valid against iLEVEL -- no rework needed there."],
        ["What actually changes",
         "Only the connection method and client library: REST/OAuth2 (requests) instead of "
         "ODBC/Entra ID (pyodbc + azure-identity). The underlying data item ids, names, categories, "
         "investment ids, and values are unchanged between the two systems."],
        ["Open follow-up (not yet resolved)",
         "Each iLEVEL data item exposes 'dependents' and 'precedents' relationships that were not "
         "populated in the one sample record we inspected. These may directly encode retired-item to "
         "replacement-item chains, which would let us replace the manual RENAME_MAP guesses in the "
         "reconciliation workbook with an authoritative API-sourced mapping. Worth a targeted follow-up "
         "call against the 6 OLD items specifically."],
    ]
    add_rows(ws, rows)
    autosize(ws, [30, 100])

    # --- Tab 2: Connection & Auth ---
    ws2 = wb.create_sheet("Connection & Auth")
    ws2.append(["Aspect", "Fabric (old)", "iLEVEL (new)"])
    style_header(ws2, 3)
    rows2 = [
        ["Protocol", "ODBC (Driver 18 for SQL Server) over TDS", "HTTPS REST, JSON:API conventions"],
        ["Auth", "Azure AD service principal via ClientSecretCredential; struct-packed access "
                 "token passed through SQL_COPT_SS_ACCESS_TOKEN",
         "OAuth2 client_credentials grant: POST /token with HTTP Basic auth (client_id/client_secret), "
         "returns access_token used as Bearer token"],
        ["Token lifetime", "~1 hr (standard Azure AD token)", "~30 min"],
        ["Env vars (removed)", "FABRIC_CLIENT_ID, FABRIC_CLIENT_SECRET, FABRIC_TENANT_ID, "
                               "FABRIC_SQL_ENDPOINT, FABRIC_DATABASE", ""],
        ["Env vars (added)", "", "ILEVEL_API_BASE_URL, ILEVEL_CLIENT_ID, ILEVEL_CLIENT_SECRET"],
        ["Python libs", "pyodbc, azure-identity", "requests (pyodbc/azure-identity no longer needed "
                                                   "once Fabric scripts are retired)"],
        ["Query language", "T-SQL (SELECT ... WHERE ... via cursor)", "Query string params: "
                                                                       "fields[<resource>], filter[<field>]=op(value), sort=, page[number]/page[size]"],
        ["Pagination", "N/A -- pulled full result sets in one query", "Required -- periodicData alone "
                                                                       "is ~2.75M rows / ~550k pages at page size 5; must page through with a larger page[size] "
                                                                       "and loop until a page returns fewer rows than requested"],
    ]
    add_rows(ws2, rows2)
    autosize(ws2, [22, 55, 65])

    # --- Tab 3: Resource Mapping ---
    ws3 = wb.create_sheet("Resource Mapping")
    ws3.append(["Data concept", "Fabric source", "iLEVEL source", "Notes"])
    style_header(ws3, 4)
    rows3 = [
        ["Financial-statement line items (dimension)", "production.base_ilevel__data_items",
         "GET /dataItems", "Same ids/names/categories confirmed live; 1,683 total items, paginated "
                            "100/page recommended"],
        ["Financial data values (fact)", "production.base_ilevel__periodic_data", "GET /periodicData",
         "Same schema and values; ~2.75M rows total across all investments/scenarios; must filter "
         "client-side or via filter[investment]/filter[lastModifiedDate] (no confirmed filter[dataItem] "
         "or filter[scenario] param yet -- needs a targeted test call)"],
        ["Portfolio companies", "production.base_ilevel__assets (not used -- manual allowlist instead)",
         "GET /entities/assets", "116 assets currently returned; the manual company_sync_list allowlist "
                                  "table still governs which companies we actually pull, per your earlier decision -- "
                                  "this endpoint is not used for roster validation"],
        ["Scenarios (Actual/Budget/etc.)", "periodic_data_scenario string column, filtered to 'Actual'",
         "GET /scenarios", "Confirmed: scenario id 1 = \"Actual\" (shortName \"Act\"). periodicData rows "
                            "carry both the scenario name and a scenarioObject.id relationship"],
        ["Item-to-asset applicability", "production.base_ilevel__data_item_relationships (dead end -- "
                                         "this was entity-type applicability, not a rename map)",
         "dataItems relationships.assets / attributes.appliedToAllAssets",
         "Not needed given the manual allowlist approach; not investigated further"],
    ]
    add_rows(ws3, rows3)
    autosize(ws3, [38, 48, 30, 70])

    # --- Tab 4: periodicData field mapping ---
    ws4 = wb.create_sheet("periodicData Fields")
    ws4.append(["Fabric column", "iLEVEL field (path)", "Example value (live pull)", "Status / notes"])
    style_header(ws4, 4)
    rows4 = [
        ["periodic_data_value", "attributes.value", "18695895.7968", "direct match"],
        ["periodic_data_data_item", "attributes.dataItem", "Long Term Loans", "direct match (name string)"],
        ["periodic_data_data_item_id", "relationships.dataItemObject.data.id", "10728",
         "direct match -- confirmed same id space as Fabric"],
        ["periodic_data_investment", "attributes.investment", "WatchMojo", "direct match (name string)"],
        ["periodic_data_investment_id", "relationships.investmentObject.data.id", "492",
         "likely matches Fabric's periodic_data_investment_id 1:1 (same underlying asset ids), but not "
         "yet cross-checked row-for-row -- verify before relying on it for the 52-company allowlist join"],
        ["periodic_data_scenario", "attributes.scenario", "Actual", "direct match (name string)"],
        ["(no Fabric column seen)", "relationships.scenarioObject.data.id", "1",
         "new -- numeric scenario id, id 1 = Actual"],
        ["periodic_data_period_end", "attributes.periodEnd", "2026-05-31", "direct match"],
        ["(no Fabric column seen)", "attributes.periodLength / periodLengthCode", "Month",
         "new field available; not previously queried from Fabric but likely exists there too"],
        ["(no Fabric column seen)", "attributes.currency", "USD", "new -- was implicit/assumed USD before"],
        ["(no Fabric column seen)", "attributes.asOfDate", "2026-07-31", "new -- reporting as-of date"],
        ["(no Fabric column seen)", "attributes.dataSource", "Excel Add-In", "new -- provenance of the entry"],
        ["data_item_last_modified_date (was per-item, not per-row)", "attributes.lastModifiedDate",
         "2026-07-20T20:03:45Z", "new -- per-row modification timestamp, more granular than Fabric's "
                                  "per-item version"],
        ["(no Fabric column seen)", "attributes.submittedBy", "REDACTED -- contains a real user email address",
         "PII -- do not persist or log this field downstream; exclude from bronze schema per org policy"],
        ["(no Fabric column seen)", "attributes.relationshipPath", "Self",
         "new -- relevant for fund look-through structures; expect \"Self\" for direct borrower financials"],
    ]
    fills4 = [None] * len(rows4)
    for i, r in enumerate(rows4):
        if r[0].startswith("(no Fabric column"):
            fills4[i] = NEW_FIELD_FILL
    fills4[-3] = PII_FILL  # submittedBy row
    add_rows(ws4, rows4, fills4)
    autosize(ws4, [42, 42, 45, 70])

    # --- Tab 5: dataItems field mapping ---
    ws5 = wb.create_sheet("dataItems Fields")
    ws5.append(["Fabric column", "iLEVEL field (path)", "Notes"])
    style_header(ws5, 3)
    rows5 = [
        ["data_item_id", "id", "direct match"],
        ["data_item_name", "attributes.name", "direct match"],
        ["data_item_category", "attributes.category", "direct match (Income Statement / Balance Sheet / "
                                                        "Cash Flow / etc.)"],
        ["is_carry_over", "attributes.isCarryover", "direct match"],
        ["is_carry_backward", "attributes.isCarrybackward", "direct match"],
        ["data_item_last_modified_date", "attributes.lastModifiedDate", "direct match"],
        ["(no Fabric column seen)", "attributes.excelTicker", "new -- short code used in the Excel add-in"],
        ["(no Fabric column seen)", "attributes.valueType / formatType", "new -- e.g. Numeric/Decimal, "
                                                                          "Numeric/Percentage"],
        ["(no Fabric column seen)", "attributes.aggregationType", "new -- e.g. \"End of Period\""],
        ["(no Fabric column seen)", "attributes.isMonetary / isPutable / isScalable / isDaily", "new flags, "
                                                                                                  "not previously pulled from Fabric"],
        ["(no Fabric column seen)", "attributes.appliedToAllAssets", "new -- whether the item applies "
                                                                      "globally or only to specific assets"],
        ["(no Fabric column seen)", "relationships.dependents / relationships.precedents",
         "UNRESOLVED -- empty in the one sample record checked so far. Worth querying directly for the "
         "6 OLD DO NOT USE item ids to see if this authoritatively encodes the retired-to-replacement "
         "mapping, instead of the name-suffix heuristic used in the reconciliation workbook"],
    ]
    fills5 = [FLAG_FILL if r[0].startswith("(no Fabric") and "dependents" in r[1] else
              (NEW_FIELD_FILL if r[0].startswith("(no Fabric") else None) for r in rows5]
    add_rows(ws5, rows5, fills5)
    autosize(ws5, [32, 45, 80])

    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")


if __name__ == "__main__":
    build()
