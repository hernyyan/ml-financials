"""
Temporary script -- appends the OLD DO NOT USE investigation as a new tab
to the existing reconciliation workbook in Downloads.
"""
import csv
import os

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "exploration")
WORKBOOK_PATH = os.path.join(os.path.expanduser("~"), "Downloads", "fabric_line_item_reconciliation.xlsx")

FLAG_NOTES = {
    10729: "Real ongoing data (2,540 rows, 41 companies) through 2025-07-31, no replacement -- needs a decision",
    10036: "Overlaps its replacement for years (both used concurrently 2016-2025) -- bronze must combine both sources",
    10724: "Real ongoing data (6,743 rows, 55 companies) through 2025-07-31, no replacement -- needs a decision",
    10725: "Real ongoing data (2,059 rows, 38 companies) through 2025-07-31, no replacement -- needs a decision",
    10033: "Zero rows ever recorded -- trivial, no migration needed",
    10063: "Zero rows ever recorded -- trivial, no migration needed",
}


def main():
    with open(os.path.join(DATA_DIR, "old_items_investigation.csv"), encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    wb = load_workbook(WORKBOOK_PATH)
    if "OLD Items Investigation" in wb.sheetnames:
        del wb["OLD Items Investigation"]
    ws = wb.create_sheet("OLD Items Investigation")

    header = [
        "old_item_id", "old_item_name", "old_row_count", "old_distinct_companies",
        "old_earliest_period_end", "old_latest_period_end",
        "replacement_item_id", "replacement_item_name",
        "replacement_row_count", "replacement_distinct_companies", "replacement_earliest_period_end",
        "notes",
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    flag_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    trivial_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for r in rows:
        old_id = int(r["old_item_id"])
        note = FLAG_NOTES.get(old_id, "")
        row = [
            r["old_item_id"], r["old_item_name"], r["old_row_count"], r["old_distinct_companies"],
            r["old_earliest_period_end"], r["old_latest_period_end"],
            r["replacement_item_id"], r["replacement_item_name"],
            r["replacement_row_count"], r["replacement_distinct_companies"], r["replacement_earliest_period_end"],
            note,
        ]
        ws.append(row)
        fill = trivial_fill if r["old_row_count"] == "0" else (flag_fill if old_id in (10729, 10036, 10724, 10725) else None)
        if fill:
            for col in range(1, len(header) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

    for col_idx in range(1, ws.max_column + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(WORKBOOK_PATH)
    print(f"Saved tab 'OLD Items Investigation' -> {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
